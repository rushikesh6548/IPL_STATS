from bs4 import BeautifulSoup
import re
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import pandas as pd
import time
from datetime import datetime
import openpyxl 

#save entire html of the page to txt file
def get_full_commentary(match_url,file_name):
    """
    Fetch full ball-by-ball commentary from ESPN Cricinfo.
    """
    options = webdriver.ChromeOptions()
    #options.add_argument("--headless")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")

    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    driver.get(match_url)
    wait = WebDriverWait(driver, 10)

    #ball_by_ball_data = []
    prev_count = 0  # Track previous entry count
    #html_content

    """while True:
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")  # Scroll down
        time.sleep(2)  # Allow time for new content to load

        try:
            #load_more_button = wait.until(EC.presence_of_element_located((By.XPATH, '//button[contains(text(), "Load More Commentary")]')))
            #driver.execute_script("arguments[0].click();", load_more_button)
            #time.sleep(2)
            load_more_selectors = [
            "//button[contains(text(), 'Load More Commentary')]",
            "//button[contains(text(), 'Load More')]",
            "//a[contains(text(), 'Load More Commentary')]",
            "//div[contains(@class, 'cb-com-more')]",
            "//span[contains(text(), 'LOAD MORE')]"
            ]
    
            button_found = False
            for selector in load_more_selectors:
                try:
                    load_more_button = wait.until(EC.element_to_be_clickable((By.XPATH, selector)))
                    driver.execute_script("arguments[0].scrollIntoView();", load_more_button)
                    time.sleep(1)
                    driver.execute_script("arguments[0].click();", load_more_button)
                    button_found = True
                    print("Clicked 'Load More' button")
                    time.sleep(2)
                    break
                except:
                    continue
            
            if not button_found:
                print("No 'Load More' button found - might be all data or different structure")
        except:
            pass  # If button is missing, continue extracting

        commentary_blocks = driver.find_elements(By.CSS_SELECTOR, ".cb-col cb-col-100 ng-scope")  # Adjust selector if necessary

        # If no new data is loaded, break the loop
        if len(commentary_blocks) == prev_count:
            print("✅ All commentary data loaded.")
            break
        prev_count = len(commentary_blocks)""" 
    
    while True:
        # Scroll down to bottom
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(2)  # Allow time for new content to load
        
        # On first iteration, try to find and click "Load More" button if it exists
        if prev_count == 0:
            try:
                load_more_selectors = [
                    "//button[contains(text(), 'Load More Commentary')]",
                    "//button[contains(text(), 'Load More')]",
                    "//a[contains(text(), 'Load More Commentary')]",
                    "//div[contains(@class, 'cb-com-more')]",
                    "//span[contains(text(), 'LOAD MORE')]"
                ]
                
                button_found = False
                for selector in load_more_selectors:
                    try:
                        load_more_button = wait.until(EC.element_to_be_clickable((By.XPATH, selector)))
                        driver.execute_script("arguments[0].scrollIntoView();", load_more_button)
                        time.sleep(1)
                        driver.execute_script("arguments[0].click();", load_more_button)
                        button_found = True
                        #print("Clicked 'Load More' button")
                        time.sleep(2)
                        break
                    except:
                        continue
                        
                if not button_found:
                    print("No initial 'Load More' button found - continuing with scrolling")
            except:
                print("Using scroll-based loading")
        
        # Get current height for comparison after scrolling
        last_height = driver.execute_script("return document.body.scrollHeight")
        
        # Scroll down again
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        
        # Wait for new content to load
        time.sleep(3)
        
        # Calculate new scroll height and compare with last scroll height
        new_height = driver.execute_script("return document.body.scrollHeight")
        
        # Find all commentary blocks with the correct selector (with dots between class names)
        commentary_blocks = driver.find_elements(By.CSS_SELECTOR, ".cb-col.cb-col-100.ng-scope")
        #print(f"Found {len(commentary_blocks)} commentary blocks")
        
        # If no new content loaded (either by count or by height), we've reached the end
        if len(commentary_blocks) == prev_count or new_height == last_height:
            print("✅ All commentary data loaded.")
            break
        
        prev_count = len(commentary_blocks)
        #print(f"Loaded {prev_count} blocks, scrolling for more...")

    #save the html content to txt file
    html_content = driver.page_source
    with open(file_name, 'w', encoding='utf-8') as file:
        file.write(str(html_content))
        
    print(f"Successfully saved HTML content to {file_name}")
    driver.quit()

    """commentary_blocks = driver.find_elements(By.CSS_SELECTOR, ".ds-text-tight-m.ds-font-regular.ds-text-ui-typo")
        overs = driver.find_elements(By.CSS_SELECTOR, ".ds-text-tight-s.ds-font-regular.ds-text-ui-typo-mid")
        players = driver.find_elements(By.CSS_SELECTOR, ".ci-ball-by-ball-commentary__player-name")

        if len(commentary_blocks) == prev_count:
            print("✅ All commentary data loaded.")
            break
        prev_count = len(commentary_blocks)

    last_over = "Unknown"
    last_batsman = "Unknown"
    last_bowler = "Unknown"

    for i in range(len(commentary_blocks)):
        commentary_text = commentary_blocks[i].text.strip() if i < len(commentary_blocks) else ""
        over = overs[i].text.strip() if i < len(overs) and overs[i].text.strip() != "" else last_over
        batsman = players[i * 2].text.strip() if (i * 2) < len(players) else last_batsman  # Even index
        bowler = players[i * 2 + 1].text.strip() if (i * 2 + 1) < len(players) else last_bowler  # Odd index

        last_over = over if over != "Unknown" else last_over
        last_batsman = batsman if batsman != "Unknown" else last_batsman
        last_bowler = bowler if bowler != "Unknown" else last_bowler

        ball_by_ball_data.append({
            "over": over,
            "batsman": batsman,
            "bowler": bowler,
            "commentary": commentary_text
        })
    
"""

    #response = requests.get(url, headers=headers)
    #response.raise_for_status()  # Raise exception for 4XX/5XX responses
        
        # Parse the HTML content with BeautifulSoup
    #soup = BeautifulSoup(response.text, 'html.parser')

    #html_content = driver.page_source

        # Save the entire HTML to a text file
    """with open(output_filename, 'w', encoding='utf-8') as file:
        file.write(str(html_content))
        
    print(f"Successfully saved HTML content to {output_filename}")
    driver.quit()"""
    #return ball_by_ball_data


#input url
match_url = input("🔗 Enter Cricbuzz match commentary URL: ").strip()

file_name = f"html_file_{datetime.now().strftime('%Y%m%d_%H')}.txt"

print("📡 Fetching ball-by-ball commentary...")

#call the function 
ball_data = get_full_commentary(match_url,file_name)

#print("🎉 Data scraping completed.")


# Load HTML from file
file_path = f"./html_file_{datetime.now().strftime('%Y%m%d_%H')}.txt"  # Change this if needed
with open(file_path, "r", encoding="utf-8") as file:
    html_content = file.read()

# Parse HTML using lxml parser
soup = BeautifulSoup(html_content, "html.parser")

# Find all divs where id starts with 'comm_'
commentary_blocks = soup.find_all("div", id=re.compile(r"^comm_"))

# Extract relevant information
ball_data = []
for block in commentary_blocks:
    over_element = block.find("div", class_="cb-mat-mnu-wrp cb-ovr-num ng-binding ng-scope")
    commentary_element = block.find("p", class_="cb-com-ln ng-binding ng-scope cb-col cb-col-90")

    if over_element and commentary_element:
        over = over_element.text.strip()
        commentary = commentary_element.text.strip()
        ball_data.append((over, commentary))


# save to excel
def write_tuples_to_excel(data, filename="output.xlsx", sheet_name="Sheet1"):
    """
    Writes a list of tuples to an Excel file.

    Args:
        data: A list of tuples, where each tuple contains two values.
        filename: The name of the Excel file to create (e.g., "output.xlsx").
        sheet_name: The name of the worksheet within the Excel file.
    """

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name

    # Write headers (optional, but good practice)
    sheet.cell(row=1, column=1, value="Value 1")
    sheet.cell(row=1, column=2, value="Value 2")

    # Write the data from the list of tuples
    for row_num, tuple_data in enumerate(data, start=2):  # Start from row 2
        value1, value2 = tuple_data
        sheet.cell(row=row_num, column=1, value=value1)
        sheet.cell(row=row_num, column=2, value=value2)

    workbook.save(filename)
    print(f"Data written to {filename}")

# --- Example Usage ---


write_tuples_to_excel(ball_data, f"cricbuzz_ball_by_ball_{datetime.now().strftime('%Y%m%d_%H')}.xlsx", "Data")

# Print first 10 records as a sample
"""for record in ball_data:
    print(record)
"""